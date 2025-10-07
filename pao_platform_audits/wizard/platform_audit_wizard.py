from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
from io import BytesIO

# usar openpyxl para .xlsx
try:
    import openpyxl
except ImportError:
    openpyxl = None

class PlatformAuditWizard(models.TransientModel):
    _name = 'pao.azz.platform.audits.wizard'
    _description = 'Wizard to Import AZZ Platform Audit'

    file = fields.Binary(string="File", required=True)
    file_name = fields.Char(string="File Name")
    sheet_name = fields.Char(string="Sheet Name")

    def action_process_file(self):
        """Procesar el archivo y devolver notificación con el resultado o errores."""
        if not self.file:
            raise UserError(_("Choose a File"))

        if openpyxl is None:
            raise UserError(_("The openpyxl library is not install."))

        data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(filename=BytesIO(data), read_only=True, data_only=True)

        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        errors = []
        created = 0
        line_no = 1

        # ejemplo: supongamos que columna A = código, B = nombre, C = categoría
        # ajustar index según tu excel (openpyxl rows start at 1)
        for row in ws.iter_rows(min_row=2):  # saltar header
            line_no += 1
            code_cell = row[0].value  # A
            name_cell = row[1].value  # B
            category_cell = row[2].value  # C

            # Normalizar / validar
            code = (str(code_cell).strip() if code_cell is not None else '')
            name = (str(name_cell).strip() if name_cell is not None else '')
            category = (str(category_cell).strip() if category_cell is not None else '')

            # ejemplo de validaciones
            if not code:
                errors.append(_("Fila %s: falta código.") % line_no)
                continue
            if not name:
                errors.append(_("Fila %s: falta nombre.") % line_no)
                continue
            if len(code) > 32:
                errors.append(_("Fila %s: código demasiado largo (>32).") % line_no)
                continue

            # ejemplo: crear o buscar categoría en model 'product.category' (ajusta a tus modelos)
            cat = None
            if category:
                cat = self.env['product.category'].search([('name', '=', category)], limit=1)
                if not cat:
                    # crear categoría si no existe
                    cat = self.env['product.category'].create({'name': category})

            # ejemplo: crear un producto-catalogo en tu modelo target (ajusta)
            # si quieres evitar duplicados:
            existing = self.env['product.product'].search([('default_code', '=', code)], limit=1)
            if existing:
                # podrías actualizar campos en lugar de crear
                existing.write({
                    'name': name,
                    'categ_id': cat.id if cat else False,
                })
            else:
                vals = {
                    'name': name,
                    'default_code': code,
                    'categ_id': cat.id if cat else False,
                }
                try:
                    self.env['product.product'].create(vals)
                    created += 1
                except Exception as e:
                    errors.append(_("Fila %s: error al crear producto: %s") % (line_no, e))

        # construir mensaje final
        if errors:
            # mostrar errores al usuario. Si son muchos, puedes crear un attachment con el log.
            body = "<b>%s</b><br/>" % _("Errores al procesar archivo")
            for err in errors[:200]:
                body += "%s<br/>" % err
            # Mostrar la notificación con el detalle
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Errores'),
                    'message': "%s. %s" % (_('Se produjeron errores'), _('Revisa las filas.')),
                    'sticky': True,
                }
            }

        # si todo bien, notificar éxito (puede ser notification simple)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación correcta'),
                'message': _('%d registros creados/actualizados') % created,
                'sticky': False,
            }
        }