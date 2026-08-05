# -*- coding: utf-8 -*-
import base64
import io
import json
import re
import xml.etree.ElementTree as ET

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

KML_NS = {'kml': 'http://www.opengis.net/kml/2.2'}

# Candidate header substrings per logical column, tried in order, tolerant of
# accent/casing/spacing variations and of different client templates (the
# standardized template uses "Nombre Sitio", "Frutos", "SUPERFICIE (HA)",
# "UBICACIÓN", "latitud", "longitud"; older files used "HUERTO",
# "PROPIETARIO", "VARIEDAD" and a single combined "COORDENADAS" column).
EXCEL_COLUMN_MARKERS = {
    'name': ['NOMBRE SITIO', 'NOMBRE', 'HUERTO'],
    'owner': ['PROPIETARIO'],
    'variety': ['FRUTOS', 'VARIEDAD'],
    'declared_ha': ['SUPERFICIE'],
    'location': ['UBICACIÓN', 'UBICACION'],
    'lat': ['LATITUD', 'LAT'],
    'lng': ['LONGITUD', 'LNG', 'LONG'],
    'coords': ['COORDENADAS'],
}

# Matches a degrees/minutes/seconds coordinate with optional hemisphere letter,
# e.g. 18°30'32.46"N, 100 34 43.5 O, -100° 34' 43.5". Degrees is the only
# required group; minutes/seconds default to 0 when absent.
DMS_HEMISPHERE_RE = re.compile(r'[NSEWOnsewo]')
DMS_NUMBER_RE = re.compile(r'-?\d+(?:[.,]\d+)?')


class PaoSitePlotImportWizard(models.TransientModel):
    _name = 'pao.site.plot.import.wizard'
    _description = 'Importar sitios desde KML o Excel'

    sale_order_id = fields.Many2one(
        comodel_name='sale.order',
        string='Cotización',
        required=True,
        default=lambda self: self.env.context.get('active_id'),
    )
    import_file = fields.Binary(string='Archivo (.kml o .xlsx)', required=True, attachment=True)
    file_name = fields.Char(string='Nombre del archivo')
    state = fields.Selection(
        selection=[('draft', 'Listo'), ('preview', 'Vista previa'), ('done', 'Importado')],
        default='draft',
    )
    preview_html = fields.Html(string='Vista previa', readonly=True)
    total_rows = fields.Integer(string='Total de filas leídas', readonly=True)
    valid_rows = fields.Integer(string='Filas válidas', readonly=True)
    error_rows = fields.Integer(string='Filas con error', readonly=True)
    parsed_rows = fields.Text()

    def _file_type(self):
        self.ensure_one()
        return 'kml' if (self.file_name or '').lower().endswith('.kml') else 'xlsx'

    # ── Preview ────────────────────────────────────────────────────────────
    def action_preview(self):
        self.ensure_one()
        file_type = self._file_type()
        if file_type == 'kml':
            rows, errors = self._parse_kml()
        else:
            if not openpyxl:
                raise UserError(_('The openpyxl library is not installed on the server.'))
            rows, errors = self._parse_excel()

        self.total_rows = len(rows) + len(errors)
        self.valid_rows = len(rows)
        self.error_rows = len(errors)
        self.parsed_rows = json.dumps(rows)
        self.preview_html = self._render_preview_html(rows, errors)
        self.state = 'preview'
        return self._reopen()

    def _render_preview_html(self, rows, errors):
        html = [
            '<div style="overflow-x:auto; max-width:100%;">'
            '<table class="table table-sm table-bordered" '
            'style="font-size:12px; white-space:nowrap; width:auto;">'
        ]
        html.append(
            '<thead class="table-dark"><tr>'
            '<th>#</th><th>Sitio</th><th>Propietario</th><th>Frutos</th>'
            '<th>Ubicación</th><th>Superficie decl. (HA)</th>'
            '<th>Latitud</th><th>Longitud</th><th>Polígono</th></tr></thead><tbody>'
        )
        for i, row in enumerate(rows[:20], 1):
            html.append(
                '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'
                '<td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                    i,
                    row.get('name', ''),
                    row.get('owner', ''),
                    row.get('variety', ''),
                    row.get('location', ''),
                    row.get('declared_ha', ''),
                    row.get('lat', ''),
                    row.get('lng', ''),
                    _('Sí') if row.get('geojson') else _('Pendiente de dibujar'),
                )
            )
        if len(rows) > 20:
            html.append(
                '<tr><td colspan="9" class="text-center text-muted">%s</td></tr>'
                % (_('… y %d filas más') % (len(rows) - 20))
            )
        html.append('</tbody></table></div>')
        if errors:
            html.append('<div class="alert alert-warning mt-2"><b>%s</b><ul>' % _('Filas con problemas:'))
            for err in errors:
                html.append('<li>%s</li>' % err)
            html.append('</ul></div>')
        return ''.join(html)

    # ── Import ─────────────────────────────────────────────────────────────
    def action_import(self):
        self.ensure_one()
        rows = json.loads(self.parsed_rows or '[]')
        if not rows:
            raise UserError(_('No valid rows were found to import.'))

        Site = self.env['pao.site.plot']
        for row in rows:
            Site.create({
                'sale_order_id': self.sale_order_id.id,
                'name': row.get('name') or _('Sin nombre'),
                'partner_owner_name': row.get('owner') or '',
                'variety': row.get('variety') or '',
                'location': row.get('location') or '',
                'declared_surface_ha': row.get('declared_ha') or 0.0,
                'center_lat': row.get('lat'),
                'center_lng': row.get('lng'),
                'geojson_polygon': json.dumps(row['geojson']) if row.get('geojson') else False,
                'source': 'kml_import' if row.get('geojson') else 'excel_import',
            })
        self.state = 'done'
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import complete'),
                'message': _('%d sites imported.') % len(rows),
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    # ── KML branch ───────────────────────────────────────────────────────────
    def _parse_kml(self):
        try:
            content = base64.b64decode(self.import_file)
            root = ET.fromstring(content)
        except ET.ParseError as e:
            raise UserError(_('Could not read the KML file: %s') % str(e))

        rows = []
        errors = []
        for i, placemark in enumerate(root.iter('{%s}Placemark' % KML_NS['kml']), 1):
            name_el = placemark.find('kml:name', KML_NS)
            name = (name_el.text or '').strip() if name_el is not None and name_el.text else _('Sitio %d') % i
            coords_el = placemark.find('.//kml:outerBoundaryIs/kml:LinearRing/kml:coordinates', KML_NS)
            if coords_el is None or not coords_el.text:
                errors.append(_('Placemark "%s": no polygon found (skipped).') % name)
                continue
            try:
                ring = []
                for pair in coords_el.text.strip().split():
                    parts = pair.split(',')
                    lng, lat = float(parts[0]), float(parts[1])
                    ring.append([lng, lat])
                rows.append({
                    'name': name,
                    'geojson': {'type': 'Polygon', 'coordinates': [ring]},
                })
            except (ValueError, IndexError) as e:
                errors.append(_('Placemark "%s": invalid coordinates (%s).') % (name, e))
        return rows, errors

    # ── Excel branch ─────────────────────────────────────────────────────────
    def _parse_excel(self):
        try:
            content = base64.b64decode(self.import_file)
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            raise UserError(_('Could not read the Excel file: %s') % str(e))
        ws = wb.active

        header_row_idx, col_index = self._find_excel_header(ws)
        if header_row_idx is None:
            raise UserError(
                _('Could not find a header row with the expected columns '
                  '(Nombre Sitio, Frutos, SUPERFICIE (HA), latitud, longitud).')
            )
        has_split_coords = 'lat' in col_index and 'lng' in col_index
        if not has_split_coords and 'coords' not in col_index:
            raise UserError(
                _('Could not find latitude/longitude columns (either separate '
                  '"latitud"/"longitud" columns, or a combined "COORDENADAS" column).')
            )

        rows = []
        errors = []
        for i, excel_row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            if not any(c is not None for c in excel_row):
                continue
            name = self._cell(excel_row, col_index.get('name'))
            if not name:
                continue
            try:
                if has_split_coords:
                    lat = self._parse_coordinate(self._raw_cell(excel_row, col_index.get('lat')))
                    lng = self._parse_coordinate(self._raw_cell(excel_row, col_index.get('lng')))
                else:
                    lat, lng = self._parse_center_point(self._cell(excel_row, col_index.get('coords')))
                rows.append({
                    'name': name,
                    'owner': self._cell(excel_row, col_index.get('owner')),
                    'variety': self._cell(excel_row, col_index.get('variety')),
                    'location': self._cell(excel_row, col_index.get('location')),
                    'declared_ha': self._cell_float(excel_row, col_index.get('declared_ha')),
                    'lat': lat,
                    'lng': lng,
                    'geojson': None,
                })
            except (ValueError, IndexError) as e:
                errors.append(_('Row %d: "%s" — invalid coordinates (%s).') % (i, name, e))
        return rows, errors

    @staticmethod
    def _find_excel_header(ws, max_scan_rows=15):
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
        ):
            normalized = [str(c).strip().upper() if c else '' for c in row]
            if any(
                marker in cell
                for cell in normalized
                for marker in EXCEL_COLUMN_MARKERS['name']
            ):
                col_index = {}
                for key, markers in EXCEL_COLUMN_MARKERS.items():
                    for marker in markers:
                        found = False
                        for idx, cell in enumerate(normalized):
                            if marker in cell:
                                col_index[key] = idx
                                found = True
                                break
                        if found:
                            break
                return row_idx, col_index
        return None, {}

    @staticmethod
    def _cell(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    @staticmethod
    def _raw_cell(row, idx):
        """Same as _cell but keeps numeric cells as-is (int/float), instead of
        stringifying them, so _parse_coordinate can tell a plain decimal
        apart from a DMS-formatted string without float()-vs-str guessing."""
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _cell_float(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return 0.0
        try:
            return float(row[idx])
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _parse_center_point(raw):
        """Parse a legacy combined 'lat lon' or 'lat,lon' string into (lat, lng) floats."""
        raw = (raw or '').strip()
        if not raw:
            return None, None
        parts = raw.split()
        if len(parts) != 2:
            parts = [p.strip() for p in raw.split(',') if p.strip()]
        if len(parts) != 2:
            raise ValueError(raw)
        return float(parts[0]), float(parts[1])

    @classmethod
    def _parse_coordinate(cls, raw):
        """Parse a single latitude/longitude cell that may be either a plain
        decimal degree value (18.509017, or "18,509017" with a decimal comma)
        or a degrees/minutes/seconds string (18°30'32.46"N, -100 34 43.5 O)."""
        if raw is None or raw == '':
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        text = str(raw).strip()
        if not text:
            return None
        try:
            return float(text.replace(',', '.'))
        except ValueError:
            pass
        return cls._dms_to_decimal(text)

    @staticmethod
    def _dms_to_decimal(text):
        hemisphere_match = DMS_HEMISPHERE_RE.search(text)
        hemisphere = hemisphere_match.group(0).upper() if hemisphere_match else None
        cleaned = DMS_HEMISPHERE_RE.sub('', text)
        numbers = [float(n.replace(',', '.')) for n in DMS_NUMBER_RE.findall(cleaned)]
        if not numbers:
            raise ValueError(text)
        negative = numbers[0] < 0
        degrees = abs(numbers[0])
        minutes = numbers[1] if len(numbers) > 1 else 0.0
        seconds = numbers[2] if len(numbers) > 2 else 0.0
        decimal = degrees + minutes / 60.0 + seconds / 3600.0
        if negative or hemisphere in ('S', 'W', 'O'):
            decimal = -decimal
        return decimal

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
