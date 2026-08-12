# -*- coding: utf-8 -*-
import base64
import io
from datetime import date, datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Columnas esperadas en el Excel (en orden, tal como el FAA-07)
COLUMNAS_REQUERIDAS = [
    'Organización/productor',   # 0
    '*Opción',                  # 1
    'Shipper',                  # 2
    '# Sitios de producción',   # 3
    '# PHU',                    # 4
    'Producto(s)',               # 5
    'Tipo de cultivo (CC/CNC/Perenne)',  # 6
    'Obligaciones mayores',     # 7
    'Obligaciones menores',     # 8
    'Sanciones/quejas',         # 9
    'Temporada de cosecha',     # 10
    'Justificación del muestreo',  # 11
    'Auditor (de la finca o del SGC)',  # 12
    'Fecha de vencimiento de certificado',  # 13
]


class EnaImportWizard(models.TransientModel):
    _name = 'ena.import.wizard'
    _description = 'Importar Productores ENA desde Excel'

    archivo_excel = fields.Binary(
        string='Archivo Excel (.xlsx)',
        required=True,
        attachment=True,
    )
    nombre_archivo = fields.Char(string='Nombre del archivo')
    anio = fields.Integer(
        string='Año natural',
        required=True,
        default=lambda self: date.today().year,
    )
    hoja = fields.Char(
        string='Nombre de la hoja',
        default='2025',
        required=True,
        help='Nombre exacto de la pestaña del Excel a importar (ej. 2025, 2026).',
    )
    fila_inicio = fields.Integer(
        string='Fila de inicio de datos',
        default=5,
        required=True,
        help='Número de fila donde comienzan los datos (sin contar encabezados). '
             'Normalmente es la fila 5 del Excel FAA-07.',
    )
    omitir_duplicados = fields.Boolean(
        string='Omitir duplicados',
        default=True,
        help='Si está activo, no importa productores que ya existan en el año seleccionado.',
    )

    # ── Resultado del preview ──────────────────────────────────────────────────
    state = fields.Selection(
        selection=[('draft', 'Listo'), ('preview', 'Vista previa'), ('done', 'Importado')],
        default='draft',
    )
    preview_html = fields.Html(
        string='Vista previa',
        readonly=True,
    )
    total_filas = fields.Integer(string='Total de filas leídas', readonly=True)
    filas_validas = fields.Integer(string='Filas válidas', readonly=True)
    filas_error = fields.Integer(string='Filas con error', readonly=True)
    log_errores = fields.Text(string='Detalle de errores', readonly=True)

    # ── Preview ────────────────────────────────────────────────────────────────
    def action_preview(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('La librería openpyxl no está instalada en el servidor.'))

        filas, errores = self._parsear_excel()
        self.total_filas = len(filas) + len(errores)
        self.filas_validas = len(filas)
        self.filas_error = len(errores)

        # Construir tabla HTML de preview (primeras 20 filas válidas)
        html = ['<table class="table table-sm table-bordered" style="font-size:12px">']
        html.append('<thead class="table-dark"><tr>'
                    '<th>#</th><th>Productor</th><th>Shipper</th>'
                    '<th>Producto</th><th>Justificación</th><th>Vencimiento</th>'
                    '</tr></thead><tbody>')
        for i, f in enumerate(filas[:20], 1):
            html.append(
                f'<tr><td>{i}</td>'
                f'<td>{f.get("nombre_productor", "")}</td>'
                f'<td>{f.get("shipper", "")}</td>'
                f'<td>{f.get("producto", "")}</td>'
                f'<td>{str(f.get("justificacion", ""))[:60]}…</td>'
                f'<td>{f.get("fecha_vencimiento", "")}</td></tr>'
            )
        if len(filas) > 20:
            html.append(f'<tr><td colspan="6" class="text-center text-muted">'
                        f'… y {len(filas) - 20} filas más</td></tr>')
        html.append('</tbody></table>')

        if errores:
            html.append('<div class="alert alert-warning mt-2"><b>Filas con problemas:</b><ul>')
            for e in errores:
                html.append(f'<li>{e}</li>')
            html.append('</ul></div>')

        self.preview_html = ''.join(html)
        self.log_errores = '\n'.join(errores) if errores else ''
        self.state = 'preview'
        return self._reopen()

    # ── Importar ───────────────────────────────────────────────────────────────
    def action_importar(self):
        self.ensure_one()
        filas, errores = self._parsear_excel()
        if not filas:
            raise UserError(_('No se encontraron filas válidas para importar.'))

        Organization = self.env['pao.globalgap.organization']
        EnaSolicitud = self.env['ena.solicitud']
        creados = 0
        omitidos = 0

        for fila in filas:
            nombre = fila.get('nombre_productor', '').strip()
            # Buscar organización por nombre (búsqueda flexible)
            org = Organization.search(
                [('name', 'ilike', nombre)], limit=1
            )
            if not org:
                # Crear una organización mínima si no existe
                org = Organization.create({'name': nombre})

            # Verificar duplicado en el año
            if self.omitir_duplicados:
                existe = EnaSolicitud.search([
                    ('organization_id', '=', org.id),
                    ('anio', '=', self.anio),
                ], limit=1)
                if existe:
                    omitidos += 1
                    continue

            vals = {
                'organization_id': org.id,
                'anio': self.anio,
                'shipper': fila.get('shipper', ''),
                'obligaciones_mayores': fila.get('obligaciones_mayores', 0.0),
                'obligaciones_menores': fila.get('obligaciones_menores', 0.0),
                'sanciones_quejas': fila.get('sanciones_quejas', 'N/A') or 'N/A',
                'justificacion': fila.get('justificacion', ''),
                'auditor_finca': fila.get('auditor_finca', ''),
                'fecha_vencimiento': fila.get('fecha_vencimiento'),
                'stage': 'candidato',
            }
            EnaSolicitud.create(vals)
            creados += 1

        self.state = 'done'
        msg = _('%d productores importados correctamente.') % creados
        if omitidos:
            msg += _(' %d omitidos por ser duplicados.') % omitidos
        if errores:
            msg += _(' %d filas con errores (ver log).') % len(errores)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación completada'),
                'message': msg,
                'type': 'success' if not errores else 'warning',
                'sticky': True,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    # ── Parser interno ─────────────────────────────────────────────────────────
    def _parsear_excel(self):
        """
        Lee el archivo Excel y devuelve (filas_validas, errores).
        filas_validas: list of dict con los datos de cada productor
        errores:       list of str con descripción de cada fila problemática
        """
        try:
            contenido = base64.b64decode(self.archivo_excel)
            wb = openpyxl.load_workbook(
                io.BytesIO(contenido), read_only=True, data_only=True
            )
        except Exception as e:
            raise UserError(_('No se pudo leer el archivo Excel: %s') % str(e))

        # Seleccionar hoja
        if self.hoja in wb.sheetnames:
            ws = wb[self.hoja]
        elif wb.active:
            ws = wb.active
        else:
            raise UserError(_('No se encontró la hoja "%s" en el archivo.') % self.hoja)

        filas_validas = []
        errores = []
        fila_num = 0

        for i, row in enumerate(ws.iter_rows(min_row=self.fila_inicio, values_only=True), start=self.fila_inicio):
            fila_num = i
            # Omitir filas completamente vacías
            if not any(c for c in row if c is not None):
                continue

            nombre_productor = self._celda_str(row, 1)   # Columna B
            if not nombre_productor:
                continue  # Fila sin productor, saltar silenciosamente

            try:
                fecha_venc = self._parse_fecha(row, 18)  # Columna S (índice 18)
                if not fecha_venc:
                    errores.append(_(f'Fila {fila_num}: "{nombre_productor}" — Fecha de vencimiento inválida o vacía.'))
                    continue

                justificacion = self._celda_str(row, 12)  # Columna M
                if not justificacion:
                    errores.append(_(f'Fila {fila_num}: "{nombre_productor}" — Justificación vacía.'))
                    continue

                tipo_raw = (self._celda_str(row, 7) or '').lower().strip()

                filas_validas.append({
                    'nombre_productor':    nombre_productor,
                    'shipper':             self._celda_str(row, 3),
                    'obligaciones_mayores': self._celda_float(row, 8),
                    'obligaciones_menores': self._celda_float(row, 9),
                    'sanciones_quejas':    self._celda_str(row, 10) or 'N/A',
                    'justificacion':       justificacion,
                    'auditor_finca':       self._celda_str(row, 13),
                    'fecha_vencimiento':   fecha_venc,
                    'producto':            self._celda_str(row, 6),
                })
            except Exception as e:
                errores.append(_(f'Fila {fila_num}: "{nombre_productor}" — Error inesperado: {e}'))

        return filas_validas, errores

    # ── Helpers ────────────────────────────────────────────────────────────────
    @staticmethod
    def _celda_str(row, idx):
        try:
            val = row[idx]
            return str(val).strip() if val is not None else ''
        except IndexError:
            return ''

    @staticmethod
    def _celda_float(row, idx):
        try:
            val = row[idx]
            if val is None:
                return 0.0
            return float(val)
        except (IndexError, ValueError, TypeError):
            return 0.0

    @staticmethod
    def _parse_fecha(row, idx):
        try:
            val = row[idx]
            if val is None:
                return False
            if isinstance(val, (datetime,)):
                return val.date()
            if isinstance(val, date):
                return val
            # Intentar parsear string
            for fmt in ('%d-%b-%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except ValueError:
                    continue
            # Número serial de Excel
            if isinstance(val, (int, float)):
                from openpyxl.utils.datetime import from_excel
                return from_excel(val)
            return False
        except (IndexError, Exception):
            return False

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
